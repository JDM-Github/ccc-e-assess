import re
import copy
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook


def _strip_app_no(app_no: str) -> str:
    app_no = str(app_no).strip()
    match = re.search(r'A0*(\d+)$', app_no, re.IGNORECASE)
    if match:
        return match.group(1)
    return app_no.lstrip('0') or '0'


def _normalize_classification(raw: str) -> str:
    if "transfer" in str(raw).strip().lower():
        return "Transferee"
    return "Freshman"


def _adjust_formula(formula: str, row_offset: int) -> str:
    def replace_ref(m):
        return f"{m.group(1)}{int(m.group(2)) + row_offset}"
    return re.sub(r'([A-Z]+)(\d+)', replace_ref, formula)


def _copy_row(ws, src_row: int, dest_row: int):
    for col_idx in range(1, ws.max_column + 1):
        src  = ws.cell(row=src_row,  column=col_idx)
        dest = ws.cell(row=dest_row, column=col_idx)
        if src.value is not None:
            if isinstance(src.value, str) and src.value.startswith("="):
                dest.value = _adjust_formula(src.value, dest_row - src_row)
            else:
                dest.value = src.value
        if src.has_style:
            dest.font          = copy.copy(src.font)
            dest.fill          = copy.copy(src.fill)
            dest.border        = copy.copy(src.border)
            dest.alignment     = copy.copy(src.alignment)
            dest.number_format = src.number_format


def _col_map(ws, header_row: int) -> dict:
    mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is not None:
            key = str(val).strip()
            if key not in mapping:  # first occurrence wins
                mapping[key] = col_idx
    return mapping


def _rs_col_to_sheet(col_name: str) -> str | None:
    match = re.match(r'RS_([^/]+)/', str(col_name).strip(), re.IGNORECASE)
    return match.group(1).upper() if match else None


def _read_data_rows(ws, data_start_row: int) -> list[dict]:
    """Read all populated data rows from a sheet into a list of {col_idx: value} dicts."""
    rows = []
    max_row = ws.max_row
    for row_idx in range(data_start_row, max_row + 1):
        row_data = {}
        has_value = False
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Skip formula cells — we only want raw values for merge
            if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith("=")):
                row_data[col_idx] = cell.value
                has_value = True
            else:
                row_data[col_idx] = cell.value
        if has_value:
            rows.append((row_idx, row_data, ws))
    return rows


class ExcelProcessorService:

    MASTERLIST_SHEET = " MASTERLIST "
    SCORESHEET_SHEET = "SCORESHEET"

    ML_HEADER_ROW   = 1
    ML_TEMPLATE_ROW = 2
    SS_HEADER_ROW   = 2
    SS_TEMPLATE_ROW = 3

    # ── Col index of the counter column (No. / #) — 1-based ──────────
    ML_COUNTER_COL = 1   # "No."
    SS_COUNTER_COL = 1   # "#"

    @staticmethod
    def process(applicant_bytes: bytes, answerer_bytes: bytes, template_bytes: bytes) -> tuple[bytes, int, list]:
        applicant_df    = pd.read_excel(BytesIO(applicant_bytes), header=0, dtype=str)
        answerer_sheets = pd.read_excel(BytesIO(answerer_bytes),  sheet_name=None, header=0, dtype=str)

        # ── INTEGRITY CHECKS ─────────────────────────────────────────
        integrity_errors = []
        sheet_app_sets   = {}

        for sheet_name, df in answerer_sheets.items():
            raw_nos = df["Application No."].dropna().tolist()
            seen, dupes = set(), []
            for no in raw_nos:
                key = _strip_app_no(no)
                if key in seen:
                    dupes.append(no)
                seen.add(key)
            if dupes:
                integrity_errors.append(f"Sheet '{sheet_name}' has duplicate Application No(s): {dupes}")
            sheet_app_sets[sheet_name] = set(_strip_app_no(n) for n in raw_nos)

        sheet_names = list(sheet_app_sets.keys())
        ref_sheet, ref_set = sheet_names[0], sheet_app_sets[sheet_names[0]]
        for sheet_name in sheet_names[1:]:
            cur_set = sheet_app_sets[sheet_name]
            only_in_ref, only_in_cur = ref_set - cur_set, cur_set - ref_set
            if only_in_ref or only_in_cur:
                msg = f"Sheet '{sheet_name}' does not match '{ref_sheet}'."
                if only_in_ref: msg += f" Missing: {sorted(only_in_ref)}."
                if only_in_cur: msg += f" Extra: {sorted(only_in_cur)}."
                integrity_errors.append(msg)

        if integrity_errors:
            raise ValueError("INTEGRITY_ERROR: " + " | ".join(integrity_errors))

        # ── BUILD LOOKUPS ─────────────────────────────────────────────
        applicant_lookup = {}
        for _, row in applicant_df.iterrows():
            applicant_lookup[_strip_app_no(row["Application No."])] = row

        score_lookups = {}
        for sheet_name, df in answerer_sheets.items():
            lookup = {}
            for _, row in df.iterrows():
                key = _strip_app_no(row["Application No."])
                lookup[key] = str(row["Correct"]).strip() if pd.notna(row.get("Correct")) else ""
            score_lookups[sheet_name.upper()] = lookup

        answerer_app_nos = next(iter(answerer_sheets.values()))["Application No."].dropna().tolist()

        # ── LOAD TEMPLATE ─────────────────────────────────────────────
        wb = load_workbook(BytesIO(template_bytes))
        ml = wb[ExcelProcessorService.MASTERLIST_SHEET]
        ss = wb[ExcelProcessorService.SCORESHEET_SHEET]

        ml_cols = _col_map(ml, ExcelProcessorService.ML_HEADER_ROW)
        ss_cols = _col_map(ss, ExcelProcessorService.SS_HEADER_ROW)

        # ── POPULATE ──────────────────────────────────────────────────
        counter, skipped = 0, []

        for raw_no in answerer_app_nos:
            key = _strip_app_no(raw_no)
            if key not in applicant_lookup:
                skipped.append(raw_no)
                continue

            counter += 1
            app = applicant_lookup[key]

            ml_dest = ExcelProcessorService.ML_TEMPLATE_ROW + counter - 1
            _copy_row(ml, ExcelProcessorService.ML_TEMPLATE_ROW, ml_dest)

            def ml_set(col_name, value, _dest=ml_dest):
                if col_name in ml_cols:
                    ml.cell(row=_dest, column=ml_cols[col_name]).value = value

            ml_set("No.",                                counter)
            ml_set("Application No.",                    app["Application No."])
            ml_set("Last Name",                          app["LAST NAME"])
            ml_set("First Name",                         app["FIRST NAME"])
            ml_set("Middle Name",                        app["MIDDLE NAME"])
            ml_set("Program Applied for: first choice",  app["FIRST PRIORITY PROGRAM"])
            ml_set("Program Applied for: second choice", app["SECOND PRIORITY PROGRAM"])
            ml_set("Classification (New or transferee)", _normalize_classification(app["Classification"]))
            ml_set("Email Address",                      app["Email Address"])
            ml_set("Phone Number",                       app["CONTACT NUMBER"])

            ss_dest = ExcelProcessorService.SS_TEMPLATE_ROW + counter - 1
            _copy_row(ss, ExcelProcessorService.SS_TEMPLATE_ROW, ss_dest)

            def ss_set(col_name, value, _dest=ss_dest):
                if col_name in ss_cols:
                    ss.cell(row=_dest, column=ss_cols[col_name]).value = value

            ss_set("#",                                   counter)
            ss_set("Application number",                  app["Application No."])
            ss_set("Last Name",                           app["LAST NAME"])
            ss_set("Given Name",                          app["FIRST NAME"])
            ss_set("Middle Name",                         app["MIDDLE NAME"])
            ss_set("Program Applied for: First choice",   app["FIRST PRIORITY PROGRAM"])
            ss_set("Program Applied for: Second Choice",  app["SECOND PRIORITY PROGRAM"])
            ss_set("Email Address",                       app["Email Address"])
            ss_set("Phone Number",                        app["CONTACT NUMBER"])
            ss_set("Classification",                      _normalize_classification(app["Classification"]))

            for col_name, col_idx in ss_cols.items():
                target_sheet = _rs_col_to_sheet(col_name)
                if target_sheet is None or target_sheet not in score_lookups:
                    continue
                ss.cell(row=ss_dest, column=col_idx).value = score_lookups[target_sheet].get(key, "")

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read(), counter, skipped

    @staticmethod
    def merge(files_bytes: list[bytes]) -> bytes:
        """
        Merge multiple generated output xlsx files into one.
        Uses the first file as the base (preserves headers, styles, formulas in template rows).
        Stacks data rows from all files sequentially, renumbering No. / # continuously.
        """
        if len(files_bytes) < 2:
            raise ValueError("At least 2 files are required to merge.")

        # Load all workbooks
        workbooks = [load_workbook(BytesIO(b)) for b in files_bytes]

        base_wb = workbooks[0]
        base_ml = base_wb[ExcelProcessorService.MASTERLIST_SHEET]
        base_ss = base_wb[ExcelProcessorService.SCORESHEET_SHEET]

        ML_DATA_START = ExcelProcessorService.ML_TEMPLATE_ROW  # row 2
        SS_DATA_START = ExcelProcessorService.SS_TEMPLATE_ROW  # row 3

        # Find counter col indices from base
        ml_cols = _col_map(base_ml, ExcelProcessorService.ML_HEADER_ROW)
        ss_cols = _col_map(base_ss, ExcelProcessorService.SS_HEADER_ROW)
        ml_counter_col = ml_cols.get("No.",  ExcelProcessorService.ML_COUNTER_COL)
        ss_counter_col = ss_cols.get("#",    ExcelProcessorService.SS_COUNTER_COL)

        # Collect all data rows from all files (skip base — it already has its rows)
        # We need to know how many rows the base already has
        def last_data_row(ws, data_start: int) -> int:
            last = data_start - 1
            for row_idx in range(data_start, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        last = row_idx
                        break
            return last

        ml_last = last_data_row(base_ml, ML_DATA_START)
        ss_last = last_data_row(base_ss, SS_DATA_START)

        ml_next_dest = ml_last + 1
        ss_next_dest = ss_last + 1

        # Running counter starts from how many rows base already has
        counter = ml_last - ML_DATA_START + 1

        for wb in workbooks[1:]:
            src_ml = wb[ExcelProcessorService.MASTERLIST_SHEET]
            src_ss = wb[ExcelProcessorService.SCORESHEET_SHEET]

            src_ml_last = last_data_row(src_ml, ML_DATA_START)
            src_ss_last = last_data_row(src_ss, SS_DATA_START)

            # Copy masterlist rows
            for src_row in range(ML_DATA_START, src_ml_last + 1):
                # Check if row actually has data
                has_data = any(
                    src_ml.cell(row=src_row, column=c).value is not None
                    for c in range(1, src_ml.max_column + 1)
                )
                if not has_data:
                    continue

                counter += 1
                _copy_row(src_ml, src_row, ml_next_dest) if False else None

                # Manual cell-by-cell copy (cross-workbook — _copy_row works within same wb)
                for col_idx in range(1, src_ml.max_column + 1):
                    src_cell  = src_ml.cell(row=src_row,    column=col_idx)
                    dest_cell = base_ml.cell(row=ml_next_dest, column=col_idx)
                    # Skip formula cells — paste raw values only
                    if isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                        dest_cell.value = None
                    else:
                        dest_cell.value = src_cell.value
                    if src_cell.has_style:
                        dest_cell.font          = copy.copy(src_cell.font)
                        dest_cell.fill          = copy.copy(src_cell.fill)
                        dest_cell.border        = copy.copy(src_cell.border)
                        dest_cell.alignment     = copy.copy(src_cell.alignment)
                        dest_cell.number_format = src_cell.number_format

                # Renumber
                base_ml.cell(row=ml_next_dest, column=ml_counter_col).value = counter
                ml_next_dest += 1

            # Copy scoresheet rows (same pattern, separate counter tracking for ss)
            ss_counter = ss_next_dest - SS_DATA_START  # how many ss rows we've written so far
            for src_row in range(SS_DATA_START, src_ss_last + 1):
                has_data = any(
                    src_ss.cell(row=src_row, column=c).value is not None
                    for c in range(1, src_ss.max_column + 1)
                )
                if not has_data:
                    continue

                ss_counter += 1
                for col_idx in range(1, src_ss.max_column + 1):
                    src_cell  = src_ss.cell(row=src_row,    column=col_idx)
                    dest_cell = base_ss.cell(row=ss_next_dest, column=col_idx)
                    if isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                        dest_cell.value = None
                    else:
                        dest_cell.value = src_cell.value
                    if src_cell.has_style:
                        dest_cell.font          = copy.copy(src_cell.font)
                        dest_cell.fill          = copy.copy(src_cell.fill)
                        dest_cell.border        = copy.copy(src_cell.border)
                        dest_cell.alignment     = copy.copy(src_cell.alignment)
                        dest_cell.number_format = src_cell.number_format

                base_ss.cell(row=ss_next_dest, column=ss_counter_col).value = ss_counter
                ss_next_dest += 1

        # Renumber the base file rows too (they were written correctly already,
        # but rewrite to be safe in case of any offset issues)
        for i, row_idx in enumerate(range(ML_DATA_START, ml_next_dest), start=1):
            base_ml.cell(row=row_idx, column=ml_counter_col).value = i
        for i, row_idx in enumerate(range(SS_DATA_START, ss_next_dest), start=1):
            base_ss.cell(row=row_idx, column=ss_counter_col).value = i

        out = BytesIO()
        base_wb.save(out)
        out.seek(0)
        return out.read()