from flask import Response, request
from jdm_electron_flask import JDMBlueprint, success, error
import json
from app.core.image_processor_service import ImageProcessorService

class ImageProcessorBlueprint(JDMBlueprint):
    def __init__(self):
        super().__init__("image_processor", __name__)

    @staticmethod
    def _get_image_bytes() -> bytes | None:
        f = request.files.get("file")
        if not f: return None
        raw = f.read()
        filename = (f.filename or "").lower()
        is_pdf = filename.endswith(".pdf") or f.content_type == "application/pdf"
        if is_pdf:
            page_num = ImageProcessorBlueprint._get_page_num()
            return ImageProcessorService.pdf_to_image_bytes(raw, page_num)
        return raw

    @staticmethod
    def _get_page_num() -> int:
        try:
            return int(request.form.get("page_num", 0))
        except (TypeError, ValueError):
            return 0

    @JDMBlueprint.post("/pipeline", auth=False, validate=None)
    def pipeline():
        image_bytes = ImageProcessorBlueprint._get_image_bytes()
        if not image_bytes:
            return error("No file uploaded (field: 'file')", 400)
        try:
            result = ImageProcessorService.run_full_pipeline(image_bytes)
            return success(result, f"Pipeline completed in {result['total_elapsed']}s")
        except Exception as e:
            return error(str(e), 500)

    @JDMBlueprint.post("/pipeline_no_annotation", auth=False, validate=None)
    def pipeline_no_annotation():
        image_bytes = ImageProcessorBlueprint._get_image_bytes()
        if not image_bytes:
            return error("No file uploaded (field: 'file')", 400)
        try:
            result = ImageProcessorService.run_full_pipeline_no_annotation(image_bytes)
            return success(result, f"Pipeline completed in {result['total_elapsed']}s")
        except Exception as e:
            return error(str(e), 500)

    @JDMBlueprint.post("/pipeline/batch/stream", auth=False, validate=None)
    def pipeline_batch_stream():
        f = request.files.get("file")
        if not f:
            return error("No file uploaded", 400)

        raw = f.read()
        filename = (f.filename or "").lower()
        is_pdf = filename.endswith(".pdf") or f.content_type == "application/pdf"

        try:
            start_page = max(1, int(request.form.get("start_page", 1)))
        except (TypeError, ValueError):
            start_page = 1

        try:
            upto = int(request.form.get("upto", 0))
        except (TypeError, ValueError):
            upto = 0

        try:
            max_pages = int(request.form.get("max_pages", 0))
        except (TypeError, ValueError):
            max_pages = 0

        try:
            max_workers = int(request.form.get("max_workers", 4))
        except (TypeError, ValueError):
            max_workers = 4

        def generate():
            try:
                if not is_pdf:
                    yield f"data: {json.dumps({'type': 'total', 'total_pages': 1})}\n\n"
                    result = ImageProcessorService.run_full_pipeline(raw)
                    yield f"data: {json.dumps({'type': 'page', 'page_num': 0, 'result': result['result'], 'answers': result['answers'], 'quads': result['quads'], 'box_meta': result['box_meta'], 'elapsed': result['total_elapsed'], 'error': None})}\n\n"
                    yield f"data: {json.dumps({'type': 'done'})}\n\n"
                    return

                import fitz
                doc = fitz.open(stream=raw, filetype="pdf")
                total_pages = len(doc)
                doc.close()

                start_idx = min(max(start_page - 1, 0), total_pages - 1)
                if upto > 0:
                    end_idx = min(upto - 1, total_pages - 1)
                elif max_pages > 0:
                    end_idx = min(start_idx + max_pages - 1, total_pages - 1)
                else:
                    end_idx = total_pages - 1

                page_range = range(start_idx, end_idx + 1)
                limit = len(page_range)

                yield f"data: {json.dumps({'type': 'total', 'total_pages': limit})}\n\n"

                def process_page(page_num):
                    try:
                        image_bytes = ImageProcessorService.pdf_to_image_bytes(raw, page_num)
                        result = ImageProcessorService.run_full_pipeline(image_bytes)
                        return {
                            'type': 'page',
                            'page_num': page_num,
                            'result': result['result'],
                            'answers': result['answers'],
                            'quads': result['quads'],
                            'box_meta': result['box_meta'],
                            'elapsed': result['total_elapsed'],
                            'error': None
                        }
                    except Exception as e:
                        return {
                            'type': 'page',
                            'page_num': page_num,
                            'result': None,
                            'answers': None,
                            'quads': None,
                            'box_meta': None,
                            'elapsed': None,
                            'error': str(e)
                        }

                from concurrent.futures import ThreadPoolExecutor, as_completed
                with ThreadPoolExecutor(max_workers=max_workers) as pool:
                    futures = {pool.submit(process_page, p): p for p in page_range}
                    for future in as_completed(futures):
                        yield f"data: {json.dumps(future.result())}\n\n"

                yield f"data: {json.dumps({'type': 'done'})}\n\n"

            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

        return Response(generate(), mimetype="text/event-stream",
                        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})

    @JDMBlueprint.post("/export", auth=False)
    def export_to_excel():
        import io
        import re
        import openpyxl
        import base64
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from app.core.box_config_service import AnswerQuizService

        body = request.get_json(force=True, silent=True) or {}
        pages: list = body.get("pages", [])

        if not pages:
            return error("No pages provided", 400)

        try:
            aq = AnswerQuizService(app_name="E-Assess")
            answer_quiz = aq.get_all()

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # ── Styles ────────────────────────────────────────────────────────
            HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
            HEADER_FILL    = PatternFill("solid", fgColor="1E1E2E")
            SCORE_HDR_FILL = PatternFill("solid", fgColor="2D3561")
            CENTER         = Alignment(horizontal="center", vertical="center")
            LEFT           = Alignment(horizontal="left",   vertical="center")

            GREEN_FILL     = PatternFill("solid", fgColor="C6EFCE")
            RED_FILL       = PatternFill("solid", fgColor="FFC7CE")
            DARK_RED_FILL  = PatternFill("solid", fgColor="FF0000")  # multiple answers
            YELLOW_FILL    = PatternFill("solid", fgColor="FFEB9C")

            GREEN_FONT     = Font(bold=True, color="276221", size=9)
            RED_FONT       = Font(bold=True, color="9C0006", size=9)
            DARK_RED_FONT  = Font(bold=True, color="FFFFFF", size=9)
            YELLOW_FONT    = Font(bold=True, color="9C6500", size=9)
            SCORE_FONT     = Font(bold=True, size=10)

            THIN = Side(style="thin", color="D0D0D0")
            THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

            CORRECT_SCORE_FILL = PatternFill("solid", fgColor="E8F5E9")
            WRONG_SCORE_FILL   = PatternFill("solid", fgColor="FFEBEE")

            def make_acronym(title: str) -> str:
                """
                'Verbal Reasoning' → 'VR'
                'Space Relations'  → 'SR'
                '2342'             → '2342'   (pure numeric, no change)
                '234-234'          → '234-234' (no alpha words, no change)
                """
                alpha_words = re.findall(r"[A-Za-z]+", title)
                # No alphabetic words at all → return title unchanged
                if not alpha_words:
                    return title
                # All tokens are digits → return title unchanged
                all_tokens = re.findall(r"[A-Za-z0-9]+", title)
                if all(tok.isdigit() for tok in all_tokens):
                    return title
                return "".join(w[0].upper() for w in alpha_words) or title[:4]

            def ensure_unique_sheet_name(wb, name: str) -> str:
                existing = {s.title for s in wb.worksheets}
                if name not in existing:
                    return name
                i = 2
                while f"{name}{i}" in existing:
                    i += 1
                return f"{name}{i}"

            def style_header_row(ws, col_count: int, score_start_col: int | None = None):
                for col in range(1, col_count + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font      = HEADER_FONT
                    cell.fill      = SCORE_HDR_FILL if score_start_col and col >= score_start_col else HEADER_FILL
                    cell.alignment = CENTER
                    cell.border    = THIN_BORDER

            def resolve_answer(
                answers_grid: list[list[bool]],
                columns: list[str],
                row_idx: int,
                is_combined: bool,
                check_by_col: bool,
            ) -> tuple[str, bool]:
                """Returns (value_str, is_multiple)."""
                if check_by_col:
                    checked = [
                        columns[c]
                        for c, col_bools in enumerate(answers_grid)
                        if row_idx < len(col_bools) and col_bools[row_idx]
                    ]
                else:
                    row = answers_grid[row_idx] if row_idx < len(answers_grid) else []
                    checked = [columns[c] for c, v in enumerate(row) if v and c < len(columns)]

                if not checked:
                    return "", False
                if is_combined:
                    return "".join(checked), False
                return ",".join(checked), len(checked) > 1

            def get_answer_style(val: str, is_multiple: bool, key_answer: str | None):
                """Returns (fill, font) for an answerer cell."""
                if key_answer is None:
                    return None, None
                if is_multiple:
                    return DARK_RED_FILL, DARK_RED_FONT
                if not val:
                    return YELLOW_FILL, YELLOW_FONT
                if val == key_answer:
                    return GREEN_FILL, GREEN_FONT
                return RED_FILL, RED_FONT

            sheet_map: dict[str, any]  = {}
            row_ptr:   dict[str, int]  = {}
            # track score cols per sheet: { sheet_title: (correct_col, wrong_col) }
            score_cols: dict[str, tuple[int, int]] = {}

            for page in pages:
                box_metas: list[dict] = page.get("box_meta", [])
                answers:   list       = page.get("answers", [])

                shared_boxes = [
                    (bm, answers[i] if i < len(answers) else [])
                    for i, bm in enumerate(box_metas)
                    if not bm.get("has_own_sheet", False)
                ]

                def resolve_single_box_value(bm: dict, ag: list) -> str:
                    """Resolve one shared box's own grid into a single display
                    string, honoring its own is_combined/check_by_col settings.
                    Used both for standalone boxes and as a building block when
                    several boxes are merged via combined_title."""
                    cols_labels  = bm.get("columns", [])
                    is_combined  = bm.get("is_combined", False)
                    check_by_col = bm.get("check_by_col", False)

                    if not is_combined:
                        return ""

                    n_rows = bm.get("grid_rows", len(ag))
                    if check_by_col:
                        checked = []
                        for c, col_bools in enumerate(ag):
                            for r, v in enumerate(col_bools):
                                if v and r < len(cols_labels):
                                    checked.append((c, cols_labels[r]))
                        checked.sort(key=lambda x: x[0])
                        return "".join(v for _, v in checked)

                    value = ""
                    for r in range(n_rows):
                        row = ag[r] if r < len(ag) else []
                        for c, v in enumerate(row):
                            if v and c < len(cols_labels):
                                value += cols_labels[c]
                    return value

                def get_prefix_values(shared_boxes) -> list[tuple[str, str]]:
                    result = []
                    processed_combined_titles: set[str] = set()
                    processed_groups: set[str] = set()

                    for bm, ag in shared_boxes:
                        combined_title = bm.get("combined_title")
                        group = bm.get("group")

                        # combined_title takes priority: every box sharing the same
                        # combined_title is merged into a single column, joined with "-"
                        # so MM=06, DD=16, YYYY=2004 → "06-16-2004"
                        if combined_title is not None:
                            if combined_title in processed_combined_titles:
                                continue
                            processed_combined_titles.add(combined_title)

                            parts = [
                                resolve_single_box_value(cbm, cag)
                                for cbm, cag in shared_boxes
                                if cbm.get("combined_title") == combined_title
                            ]
                            merged_value = "-".join(parts)
                            result.append((combined_title, merged_value))
                            continue

                        if group is not None:
                            if group in processed_groups:
                                continue
                            processed_groups.add(group)

                            all_checked: list[str] = []
                            for gbm, gag in shared_boxes:
                                if gbm.get("combined_title") is not None:
                                    continue
                                if gbm.get("group") != group:
                                    continue
                                gcols = gbm.get("columns", [])
                                if gbm.get("check_by_col", False):
                                    for col_idx, col_bools in enumerate(gag):
                                        for row_idx, v in enumerate(col_bools):
                                            if v and row_idx < len(gcols):
                                                all_checked.append(gcols[row_idx])
                                else:
                                    for row in gag:
                                        for c_idx, v in enumerate(row):
                                            if v and c_idx < len(gcols):
                                                all_checked.append(gcols[c_idx])

                            result.append((bm.get("title", group), ",".join(all_checked)))

                        else:
                            result.append((bm.get("title", ""), resolve_single_box_value(bm, ag)))

                    return result

                prefix_values = get_prefix_values(shared_boxes)

                own_boxes = [
                    (i, bm, answers[i] if i < len(answers) else [])
                    for i, bm in enumerate(box_metas)
                    if bm.get("has_own_sheet", False)
                ]

                # ── Group own_boxes by their sheet key ────────────────────────
                # Sheet key rules:
                #   - has_own_sheet + combined_title → all boxes with the same
                #     combined_title land on ONE sheet named after the
                #     combined_title acronym, questions written sequentially,
                #     single Correct/Wrong pair tallying all of them.
                #   - has_own_sheet + no combined_title → one sheet per box
                #     (original behaviour, keyed by the box title acronym).
                #
                # We build an ordered dict: sheet_key → list of (idx, bm, ag)
                from collections import OrderedDict
                own_sheet_groups: OrderedDict[str, list] = OrderedDict()

                for _idx, bm, ag in own_boxes:
                    ct = bm.get("combined_title")
                    if ct:
                        # sheet key = acronym of the combined_title
                        skey = make_acronym(ct)
                    else:
                        # sheet key = acronym of the box's own title (original)
                        skey = make_acronym(bm.get("title", f"Box_{_idx}"))

                    if skey not in own_sheet_groups:
                        own_sheet_groups[skey] = []
                    own_sheet_groups[skey].append((_idx, bm, ag))

                # ── Write one row per sheet-group per page ────────────────────
                for skey, group_boxes in own_sheet_groups.items():

                    # Collect metadata across all boxes in this group
                    # is_answerer / key_answers: use first box that declares it
                    group_is_answerer = any(bm.get("is_answerer", False) for _, bm, _ in group_boxes)

                    # Build the merged key_answers list in box order
                    group_key_answers: list[str | None] | None = [] if group_is_answerer else None
                    if group_is_answerer:
                        for _, bm, _ in group_boxes:
                            full_title  = bm.get("title", "")
                            box_keys    = answer_quiz.get(full_title) or []
                            grid_rows_b = bm.get("grid_rows", 0)
                            # pad / trim to grid_rows so offsets are correct
                            for qi in range(grid_rows_b):
                                group_key_answers.append(box_keys[qi] if qi < len(box_keys) else None)

                    # ── Create sheet on first encounter ───────────────────────
                    if skey not in sheet_map:
                        safe_name = ensure_unique_sheet_name(wb, skey)
                        ws = wb.create_sheet(title=safe_name)
                        sheet_map[skey] = ws

                        col_cursor = 1

                        # prefix columns
                        for ptitle, _ in prefix_values:
                            ws.cell(row=1, column=col_cursor, value=ptitle)
                            col_cursor += 1

                        # question columns — sequential across all boxes in group
                        q_global = 1
                        for _, bm, ag in group_boxes:
                            grid_rows_b = bm.get("grid_rows", len(ag))
                            for _ in range(grid_rows_b):
                                ws.cell(row=1, column=col_cursor, value=f"Q{q_global}")
                                col_cursor += 1
                                q_global += 1

                        # single Correct / Wrong pair for the whole group
                        if group_is_answerer:
                            correct_col = col_cursor
                            wrong_col   = col_cursor + 1
                            ws.cell(row=1, column=correct_col, value="Correct")
                            ws.cell(row=1, column=wrong_col,   value="Wrong")
                            score_cols[skey] = (correct_col, wrong_col)
                            col_cursor += 2
                        else:
                            score_cols[skey] = (None, None)

                        style_header_row(
                            ws, col_cursor - 1,
                            score_start_col=score_cols[skey][0]
                        )

                        # freeze top row + prefix columns
                        freeze_col = len(prefix_values) + 1
                        ws.freeze_panes = ws.cell(row=2, column=freeze_col)

                        # column widths
                        for c in range(1, len(prefix_values) + 1):
                            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 14
                        for c in range(len(prefix_values) + 1, col_cursor):
                            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 6
                        if group_is_answerer:
                            ws.column_dimensions[openpyxl.utils.get_column_letter(correct_col)].width = 9
                            ws.column_dimensions[openpyxl.utils.get_column_letter(wrong_col)].width = 9

                        row_ptr[skey] = 2

                    # ── Write one data row for this page ──────────────────────
                    ws         = sheet_map[skey]
                    write_row  = row_ptr[skey]
                    col_cursor = 1
                    correct_count = 0
                    wrong_count   = 0
                    key_offset    = 0   # running index into group_key_answers

                    # prefix values
                    for _, pval in prefix_values:
                        cell = ws.cell(row=write_row, column=col_cursor, value=pval)
                        cell.alignment = LEFT
                        cell.border    = THIN_BORDER
                        col_cursor += 1

                    # answer cells — one box at a time, columns are sequential
                    for _, bm, ag in group_boxes:
                        cols_labels  = bm.get("columns", [])
                        grid_rows_b  = bm.get("grid_rows", len(ag))
                        is_combined  = bm.get("is_combined", False)
                        check_by_col = bm.get("check_by_col", False)

                        for q_idx in range(grid_rows_b):
                            val, is_multiple = resolve_answer(
                                ag, cols_labels, q_idx, is_combined, check_by_col
                            )
                            cell = ws.cell(row=write_row, column=col_cursor, value=val)
                            cell.alignment = CENTER
                            cell.border    = THIN_BORDER

                            if group_is_answerer:
                                key_answer = (
                                    group_key_answers[key_offset]
                                    if group_key_answers and key_offset < len(group_key_answers)
                                    else None
                                )
                                fill, font = get_answer_style(val, is_multiple, key_answer)
                                if fill:
                                    cell.fill = fill
                                if font:
                                    cell.font = font

                                if not is_multiple and val and key_answer:
                                    if val == key_answer:
                                        correct_count += 1
                                    else:
                                        wrong_count += 1
                                elif not val:
                                    wrong_count += 1
                                elif is_multiple:
                                    wrong_count += 1

                            col_cursor += 1
                            key_offset += 1

                    # score columns
                    c_col, w_col = score_cols[skey]
                    if c_col:
                        cc = ws.cell(row=write_row, column=c_col, value=correct_count)
                        cc.alignment = CENTER
                        cc.border    = THIN_BORDER
                        cc.font      = Font(bold=True, color="276221", size=10)
                        cc.fill      = CORRECT_SCORE_FILL

                        wc = ws.cell(row=write_row, column=w_col, value=wrong_count)
                        wc.alignment = CENTER
                        wc.border    = THIN_BORDER
                        wc.font      = Font(bold=True, color="9C0006", size=10)
                        wc.fill      = WRONG_SCORE_FILL

                    row_ptr[skey] = write_row + 1

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            b64 = base64.b64encode(buf.read()).decode("utf-8")

            return success(
                {"file": b64, "filename": "answers_export.xlsx"},
                "Export successful"
            )
        except Exception as e:
            return error(str(e), 500)